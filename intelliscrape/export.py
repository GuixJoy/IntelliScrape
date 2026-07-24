"""Data export utilities for IntelliScrape."""

from __future__ import annotations

import csv
import json
import sqlite3
from io import StringIO
from pathlib import Path
from typing import Any, Dict, List, Optional, Union


class DataExporter:
    """Export scraped data to various formats."""
    
    @staticmethod
    def to_json(
        data: Union[List[Dict], Dict],
        file: Optional[str] = None,
        indent: int = 2,
        ensure_ascii: bool = False,
    ) -> str:
        """Export data to JSON.
        
        Parameters
        ----------
        data : list or dict
            Data to export.
        file : str, optional
            File path to save. If None, returns string.
        indent : int
            JSON indentation.
        ensure_ascii : bool
            Escape non-ASCII characters.
            
        Returns
        -------
        str
            JSON string.
        """
        json_str = json.dumps(data, indent=indent, ensure_ascii=ensure_ascii)
        
        if file:
            Path(file).parent.mkdir(parents=True, exist_ok=True)
            with open(file, "w", encoding="utf-8") as f:
                f.write(json_str)
        
        return json_str
    
    @staticmethod
    def to_csv(
        data: List[Dict],
        file: Optional[str] = None,
        delimiter: str = ",",
        encoding: str = "utf-8",
    ) -> str:
        """Export data to CSV.
        
        Parameters
        ----------
        data : list of dict
            Data to export. Each dict is a row.
        file : str, optional
            File path to save. If None, returns string.
        delimiter : str
            Column delimiter.
        encoding : str
            File encoding.
            
        Returns
        -------
        str
            CSV string.
        """
        if not data:
            return ""
        
        # Get all unique keys
        fieldnames = list(dict.fromkeys(key for row in data for key in row.keys()))
        
        output = StringIO()
        writer = csv.DictWriter(output, fieldnames=fieldnames, delimiter=delimiter)
        writer.writeheader()
        writer.writerows(data)
        
        csv_str = output.getvalue()
        
        if file:
            Path(file).parent.mkdir(parents=True, exist_ok=True)
            with open(file, "w", encoding=encoding, newline="") as f:
                f.write(csv_str)
        
        return csv_str
    
    @staticmethod
    def to_excel(
        data: List[Dict],
        file: str,
        sheet_name: str = "Sheet1",
    ) -> None:
        """Export data to Excel (XLSX).
        
        Parameters
        ----------
        data : list of dict
            Data to export.
        file : str
            File path to save.
        sheet_name : str
            Excel sheet name.
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl required for Excel export: pip install openpyxl")
        
        if not data:
            return
        
        # Get all unique keys
        fieldnames = list(dict.fromkeys(key for row in data for key in row.keys()))
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Write header
        for col, field in enumerate(fieldnames, 1):
            ws.cell(row=1, column=col, value=field)
        
        # Write data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, field in enumerate(fieldnames, 1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(field, ""))
        
        Path(file).parent.mkdir(parents=True, exist_ok=True)
        wb.save(file)
    
    @staticmethod
    def to_sqlite(
        data: List[Dict],
        file: str,
        table_name: str = "scraped_data",
    ) -> None:
        """Export data to SQLite database.
        
        Parameters
        ----------
        data : list of dict
            Data to export.
        file : str
            Database file path.
        table_name : str
            Table name.
        """
        if not data:
            return
        
        # Get all unique keys
        fieldnames = list(dict.fromkeys(key for row in data for key in row.keys()))
        
        conn = sqlite3.connect(file)
        cursor = conn.cursor()
        
        # Create table
        columns = ", ".join([f"{field} TEXT" for field in fieldnames])
        cursor.execute(f"CREATE TABLE IF NOT EXISTS {table_name} ({columns})")
        
        # Insert data
        placeholders = ", ".join(["?" for _ in fieldnames])
        for row in data:
            values = [str(row.get(field, "")) for field in fieldnames]
            cursor.execute(f"INSERT INTO {table_name} VALUES ({placeholders})", values)
        
        conn.commit()
        conn.close()
    
    @staticmethod
    def to_text(
        data: Union[List[Dict], Dict, str],
        file: Optional[str] = None,
        encoding: str = "utf-8",
    ) -> str:
        """Export data to plain text.
        
        Parameters
        ----------
        data : list, dict, or str
            Data to export.
        file : str, optional
            File path to save.
        encoding : str
            File encoding.
            
        Returns
        -------
        str
            Text string.
        """
        if isinstance(data, str):
            text = data
        elif isinstance(data, dict):
            text = json.dumps(data, indent=2)
        elif isinstance(data, list):
            lines = []
            for item in data:
                if isinstance(item, dict):
                    lines.append(json.dumps(item, indent=2))
                else:
                    lines.append(str(item))
            text = "\n\n".join(lines)
        else:
            text = str(data)
        
        if file:
            Path(file).parent.mkdir(parents=True, exist_ok=True)
            with open(file, "w", encoding=encoding) as f:
                f.write(text)
        
        return text
    
    @staticmethod
    def to_markdown(
        data: List[Dict],
        file: Optional[str] = None,
    ) -> str:
        """Export data to Markdown table.
        
        Parameters
        ----------
        data : list of dict
            Data to export.
        file : str, optional
            File path to save.
            
        Returns
        -------
        str
            Markdown string.
        """
        if not data:
            return ""
        
        # Get all unique keys
        fieldnames = list(dict.fromkeys(key for row in data for key in row.keys()))
        
        # Create header
        header = "| " + " | ".join(fieldnames) + " |"
        separator = "| " + " | ".join(["---"] * len(fieldnames)) + " |"
        
        # Create rows
        rows = []
        for row in data:
            values = [str(row.get(field, "")).replace("|", "\\|") for field in fieldnames]
            rows.append("| " + " | ".join(values) + " |")
        
        md = "\n".join([header, separator] + rows)
        
        if file:
            Path(file).parent.mkdir(parents=True, exist_ok=True)
            with open(file, "w", encoding="utf-8") as f:
                f.write(md)
        
        return md
    
    @classmethod
    def export(
        cls,
        data: Union[List[Dict], Dict, str],
        format: str,
        file: Optional[str] = None,
        **kwargs,
    ) -> str:
        """Export data to specified format.
        
        Parameters
        ----------
        data : list, dict, or str
            Data to export.
        format : str
            Export format: "json", "csv", "excel", "sqlite", "text", "markdown".
        file : str, optional
            File path to save.
        **kwargs
            Additional arguments for format-specific export.
            
        Returns
        -------
        str
            Exported string (except for excel and sqlite).
        """
        format = format.lower()
        
        if format == "json":
            return cls.to_json(data, file=file, **kwargs)
        elif format == "csv":
            return cls.to_csv(data if isinstance(data, list) else [data], file=file, **kwargs)
        elif format == "excel":
            cls.to_excel(data if isinstance(data, list) else [data], file=file or "output.xlsx", **kwargs)
            return f"Saved to {file or 'output.xlsx'}"
        elif format == "sqlite":
            cls.to_sqlite(data if isinstance(data, list) else [data], file=file or "output.db", **kwargs)
            return f"Saved to {file or 'output.db'}"
        elif format == "text" or format == "txt":
            return cls.to_text(data, file=file, **kwargs)
        elif format == "markdown" or format == "md":
            return cls.to_markdown(data if isinstance(data, list) else [data], file=file, **kwargs)
        else:
            raise ValueError(f"Unknown format: {format}")
